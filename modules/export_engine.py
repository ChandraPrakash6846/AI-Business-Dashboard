# AI Business Dashboard - Fully Updated & Verified
import os
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def export_to_excel(df, health_report, kpis, insights):
    """
    Generates a formatted multi-tab Excel workbook containing:
    - Tab 1: Executive Summary & KPIs
    - Tab 2: Cleaned Dataset
    - Tab 3: AI Insights & Recommendations
    """
    output = io.BytesIO()
    wb = Workbook()
    
    # ----------------------------------------------------
    # Tab 1: Executive Summary
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    sub_font = Font(name="Arial", size=11, bold=True, color="0F172A")
    regular_font = Font(name="Arial", size=10)
    
    ws_summary["A1"] = "AI BUSINESS DASHBOARD - EXECUTIVE REPORT"
    ws_summary["A1"].font = header_font
    ws_summary["A1"].fill = header_fill
    ws_summary.merge_cells("A1:E1")
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 40
    
    # KPI Section
    ws_summary["A3"] = "KEY PERFORMANCE INDICATORS"
    ws_summary["A3"].font = sub_font
    
    r = 4
    for k, v in kpis.items():
        ws_summary[f"A{r}"] = k
        ws_summary[f"B{r}"] = str(v)
        ws_summary[f"A{r}"].font = Font(bold=True)
        r += 1
        
    r += 1
    ws_summary[f"A{r}"] = "DATASET HEALTH METRICS"
    ws_summary[f"A{r}"].font = sub_font
    r += 1
    for k, v in health_report.items():
        ws_summary[f"A{r}"] = str(k).replace("_", " ").title()
        ws_summary[f"B{r}"] = str(v)
        ws_summary[f"A{r}"].font = Font(bold=True)
        r += 1
        
    # Auto-adjust column width
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # ----------------------------------------------------
    # Tab 2: Cleaned Dataset
    # ----------------------------------------------------
    ws_data = wb.create_sheet(title="Cleaned Data")
    ws_data.views.sheetView[0].showGridLines = True
    
    for row in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(row)
        
    # Style Header
    for cell in ws_data[1]:
        cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

    # ----------------------------------------------------
    # Tab 3: Insights
    # ----------------------------------------------------
    ws_insights = wb.create_sheet(title="AI Insights")
    ws_insights.views.sheetView[0].showGridLines = True
    
    ws_insights.append(["Category", "Title", "Detailed Description"])
    for cell in ws_insights[1]:
        cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        
    for item in insights:
        ws_insights.append([item.get("type", "info").upper(), item.get("title", ""), item.get("description", "").replace("**", "")])
        
    for col in ws_insights.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_insights.column_dimensions[col_letter].width = min(max(max_len + 2, 15), 60)

    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_to_pdf(df, health_report, kpis, insights):
    """
    Generates a high-quality PDF report document using ReportLab.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=12
    )
    
    heading_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#1E293B'),
        spaceBefore=14,
        spaceAfter=8
    )
    
    body_style = ParagraphStyle(
        'BodyText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#334155'),
        spaceAfter=6
    )

    elements = []
    
    # Title & Header
    elements.append(Paragraph("AI Business Dashboard - Internship Executive Report", title_style))
    elements.append(Paragraph("Developer: Chandra Prakash Choudhary • AI & Data Science Intern", body_style))
    elements.append(Paragraph("Submitted for Project Evaluation • Generated by AI Business Dashboard System", body_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#2563EB'), spaceBefore=8, spaceAfter=15))


    
    # Section 1: Executive KPIs
    elements.append(Paragraph("1. Executive Key Performance Indicators", heading_style))
    kpi_data = [["Metric", "Value"]]
    for k, v in kpis.items():
        kpi_data.append([str(k), str(v)])
        
    t_kpi = Table(kpi_data, colWidths=[240, 280])
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
    ]))
    elements.append(t_kpi)
    elements.append(Spacer(1, 15))

    # Section 2: Dataset Cleaning & Health
    elements.append(Paragraph("2. Data Health & Validation Summary", heading_style))
    health_data = [["Metric", "Count / Status"]]
    for k, v in health_report.items():
        health_data.append([str(k).replace("_", " ").title(), str(v)])
        
    t_health = Table(health_data, colWidths=[240, 280])
    t_health.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#334155')),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FFFFFF')),
    ]))
    elements.append(t_health)
    elements.append(Spacer(1, 15))

    # Section 3: AI Insights
    elements.append(Paragraph("3. AI Analytical Insights & Strategic Recommendations", heading_style))
    for item in insights:
        title = item.get('title', 'Insight')
        desc = item.get('description', '').replace('**', '')
        elements.append(Paragraph(f"• <b>{title}</b>: {desc}", body_style))
        elements.append(Spacer(1, 4))
        
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
